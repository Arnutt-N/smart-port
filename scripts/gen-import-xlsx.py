#!/usr/bin/env python3
"""
สร้างไฟล์ Excel สำหรับ HR data import tooling:
  - docs/import-template.xlsx           เทมเพลตให้ HR กรอก (header ไทย + 1 แถวตัวอย่าง)
  - backend/tests/fixtures/import-sample.xlsx   ข้อมูลทดสอบ (ImportServiceTest)

คอลัมน์ต้องเรียงตรงกับ ImportService::SHEETS (อ่านตามตำแหน่ง ไม่ใช่ชื่อ header)
รัน: python scripts/gen-import-xlsx.py
"""
import os
from openpyxl import Workbook

HEADERS = {
    'Personnel':   ['เลขบัตรประชาชน(13หลัก)', 'ชื่อ', 'นามสกุล', 'วันบรรจุ(YYYY-MM-DD)',
                    'ระดับปัจจุบัน(M1/M2/S1/S2/K5/K3/K4/O3)', 'วันเข้าระดับ(YYYY-MM-DD)',
                    'วุฒิ(BACHELOR/MASTER/DOCTORATE)'],
    'Diverse':     ['เลขบัตรประชาชน', 'ต่างสายงาน(1/0)', 'ต่างหน่วยงาน(1/0)',
                    'ต่างพื้นที่(1/0)', 'ต่างลักษณะงาน(1/0)', 'วันครบ3ต่าง(YYYY-MM-DD)'],
    'Equivalence': ['เลขบัตรประชาชน', 'ตำแหน่งจริง', 'ประเภทเทียบ', 'จำนวนวันเทียบ',
                    'สถานะ(APPROVED)', 'วันเริ่ม(YYYY-MM-DD)', 'วันสิ้นสุด(YYYY-MM-DD)'],
    'History':     ['เลขบัตรประชาชน', 'ระดับ(M1/M2/K3/O3..)', 'วันเข้า(YYYY-MM-DD)',
                    'วันออก(YYYY-MM-DD)', 'ชื่อตำแหน่ง'],
}


def build(path, data):
    wb = Workbook()
    wb.remove(wb.active)
    for sheet, cols in HEADERS.items():
        ws = wb.create_sheet(sheet)
        ws.append(cols)
        for row in data.get(sheet, []):
            ws.append([str(c) for c in row])
        for r in range(2, ws.max_row + 1):
            ws.cell(r, 1).number_format = '@'  # citizen_id เป็น text กัน 13 หลักเพี้ยน
        for i, _ in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = 22
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print('wrote', path)


# เทมเพลต: 1 แถวตัวอย่างต่อชีต (HR ลบก่อนกรอกจริง)
build('docs/import-template.xlsx', {
    'Personnel':   [['1100100200001', 'สมชาย(ตัวอย่าง-ลบก่อนใช้)', 'ใจดี', '2010-01-01', 'K3', '2020-01-01', 'MASTER']],
    'Diverse':     [['1100100200001', '1', '1', '1', '0', '2018-01-01']],
    'Equivalence': [['1100100200001', 'ผู้เชี่ยวชาญ', 'อำนวยการ', '730', 'APPROVED', '2020-01-01', '2022-01-01']],
    'History':     [['1100100200001', 'K3', '2016-01-01', '2020-01-01', 'ชำนาญการพิเศษ']],
})

# ข้อมูลทดสอบ: 299001 = M1 (K3+3ปี+3ต่าง → 2023-01-01), 299002 = S2 (S1+1ปี → 2023-01-01)
build('backend/tests/fixtures/import-sample.xlsx', {
    'Personnel': [
        ['1100100299001', 'สมหมาย', 'ทดสอบนำเข้า', '2010-01-01', 'K3', '2020-01-01', 'MASTER'],
        ['1100100299002', 'สมศรี', 'ทดสอบนำเข้า', '2008-01-01', 'S1', '2022-01-01', 'MASTER'],
    ],
    'Diverse':     [['1100100299001', '1', '1', '1', '0', '2018-01-01']],
    'Equivalence': [],
    'History':     [],
})
