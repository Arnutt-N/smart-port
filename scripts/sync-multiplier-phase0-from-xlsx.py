#!/usr/bin/env python3
"""
Sync Phase 0 HR workbook → CSV templates used by validate-multiplier-phase0.mjs

Usage (from repo root):
  python scripts/sync-multiplier-phase0-from-xlsx.py
  python scripts/sync-multiplier-phase0-from-xlsx.py --workbook path/to/filled.xlsx

After sync, run:
  node scripts/validate-multiplier-phase0.mjs
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "docs" / "multiplier_phase0_hr_workbook.xlsx"
MASTER_CSV = ROOT / "docs" / "multiplier_phase0_master_data_template.csv"
UAT_CSV = ROOT / "docs" / "multiplier_phase0_uat_cases_template.csv"

MASTER_SHEET = "1-พื้นที่ทวีคูณ"
UAT_SHEET = "2-ตัวอย่างจริง UAT"

MASTER_KEYS = [
    "row_id",
    "province",
    "district",
    "whole_province",
    "basis_type",
    "multiplier_ratio",
    "effective_start_date",
    "effective_end_date",
    "legal_reference",
    "source_reference",
    "verified_by",
    "verified_date",
    "notes",
]

UAT_KEYS = [
    "case_id",
    "personnel_ref",
    "province",
    "district",
    "service_start_date",
    "service_end_date",
    "expected_eligible_start_date",
    "expected_eligible_end_date",
    "expected_service_days",
    "expected_eligible_days",
    "expected_effective_days",
    "expected_bonus_days",
    "expected_net_years",
    "expected_net_months",
    "expected_net_days",
    "excel_source_ref",
    "verified_by",
    "verified_date",
    "notes",
]


def cell_str(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # Excel may emit datetime as "2004-01-26 00:00:00"
    if len(s) >= 10 and s[4] == "-" and s[7] == "-" and " " in s:
        return s[:10]
    return s


def find_key_row(rows: list[list], keys: list[str]) -> int:
    for i, row in enumerate(rows):
        normalized = [cell_str(c) for c in row[: len(keys)]]
        if normalized[:3] == keys[:3]:
            return i
    raise SystemExit(f"Cannot find English key row starting with {keys[:3]}")


def normalize_placeholder(row: dict) -> dict:
    """Keep validator-friendly placeholders when HR left cells blank."""
    out = dict(row)
    for key in (
        "legal_reference",
        "source_reference",
        "verified_by",
        "verified_date",
        "excel_source_ref",
        "effective_start_date",
        "effective_end_date",
    ):
        if key in out and out[key] == "":
            out[key] = "TODO"
    # Satun must stay district-level placeholders until HR names them
    if out.get("province") == "สตูล" and out.get("district") == "":
        rid = out.get("row_id") or "?"
        out["district"] = f"TODO อำเภอ ({rid})"
        if out.get("whole_province") in ("", "yes"):
            out["whole_province"] = "no"
    return out


def sheet_to_dicts(ws, keys: list[str]) -> list[dict]:
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    key_idx = find_key_row(rows, keys)
    out = []
    for row in rows[key_idx + 1 :]:
        values = [cell_str(c) for c in row[: len(keys)]]
        if not any(values):
            continue
        # skip accidental duplicate header
        if values[0] == keys[0]:
            continue
        out.append(normalize_placeholder(dict(zip(keys, values))))
    return out


def write_csv(path: Path, keys: list[str], rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=keys, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in keys})


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync Phase 0 workbook to CSV templates")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()
    wb_path = args.workbook.resolve()
    if not wb_path.exists():
        raise SystemExit(f"Workbook not found: {wb_path}")

    wb = load_workbook(wb_path, data_only=True)
    if MASTER_SHEET not in wb.sheetnames or UAT_SHEET not in wb.sheetnames:
        raise SystemExit(
            f"Expected sheets {MASTER_SHEET!r} and {UAT_SHEET!r}; found {wb.sheetnames}"
        )

    master_rows = sheet_to_dicts(wb[MASTER_SHEET], MASTER_KEYS)
    uat_rows = sheet_to_dicts(wb[UAT_SHEET], UAT_KEYS)
    write_csv(MASTER_CSV, MASTER_KEYS, master_rows)
    write_csv(UAT_CSV, UAT_KEYS, uat_rows)
    print(f"Synced {wb_path.name}")
    print(f"  → {MASTER_CSV.relative_to(ROOT)} ({len(master_rows)} rows)")
    print(f"  → {UAT_CSV.relative_to(ROOT)} ({len(uat_rows)} rows)")
    print("Next: node scripts/validate-multiplier-phase0.mjs")


if __name__ == "__main__":
    main()
